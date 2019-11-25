from openpyxl import load_workbook
from operator import eq
import pymysql.cursors
import os

#엑셀 불러오기
load_cloth=load_workbook(os.environ['MYROUTE']+"/python/vm.xlsx", data_only=True)
cloth_excel=load_cloth['cloth']
style_excel=load_cloth['style']


#사용자정의함수 vm사용
def usedvm(bottoms1,bottoms2,bottoms3,bottoms4,bottomsstyle,top1,top2,top3,top4,topstyle) :
    #//////////////////////////////하의
    ###############################pants
    if eq(bottoms1,"pants"):
        bottoms=5
    ###############################
        if eq(bottoms2,"long"):
            bottoms+=0
        elif eq(bottoms2,"short"):
            bottoms+=40
    ###############################
        if eq(bottoms3,"bootcut"):
            bottoms+=0
        elif eq(bottoms3, "cagopants"):
            bottoms+=8
        elif eq(bottoms3, "jeans"):
            bottoms+=16
        elif eq(bottoms3, "jogger"):
            bottoms+=24
        elif eq(bottoms3, "leggings"):
            bottoms+=32
    #################################
        if eq(bottoms4,"printing"):
            bottoms+=0
        elif eq(bottoms4, "dot"):
            bottoms+=1
        elif eq(bottoms4, "leopard"):
            bottoms+=2
        elif eq(bottoms4, "race"):
            bottoms+=3
        elif eq(bottoms4, "basic"):
            bottoms+=4
        elif eq(bottoms4, "stripe"):
            bottoms+=5
        elif eq(bottoms4, "check"):
            bottoms+=6
        elif eq(bottoms4, "camouflage"):
            bottoms+=7
    ###################################@@@@@@@@@@@@@@@@
            
    ###################################skirt
    elif eq(bottoms1,"skirt") and eq(bottoms2,"skirt"):
        bottoms=85
    ###############################
        if eq(bottoms3,"accordion skirt"):
            bottoms+=0
        elif eq(bottoms3, "a-line skirt"):
            bottoms+=8
        elif eq(bottoms3, "mini skirt"):
            bottoms+=16
        elif eq(bottoms3, "tube skirt"):
            bottoms+=24
        elif eq(bottoms3, "wrap skirt"):
            bottoms+=32
    #################################
        if eq(bottoms4,"printing"):
            bottoms+=0
        elif eq(bottoms4, "dot"):
            bottoms+=1
        elif eq(bottoms4, "leopard"):
            bottoms+=2
        elif eq(bottoms4, "race"):
            bottoms+=3
        elif eq(bottoms4, "basic"):
            bottoms+=4
        elif eq(bottoms4, "stripe"):
            bottoms+=5
        elif eq(bottoms4, "check"):
            bottoms+=6
        elif eq(bottoms4, "camouflage"):
            bottoms+=7

    ###################################@@@@@@@@@@@@@@@@
    #//////////////////////////////하의


            
    #//////////////////////////////상의
    ##################################t-shirts
    if eq(top1,"t-shirts"):
        top=6
        if eq(top2,"logn sleeve"):
            top+=0
        elif eq(top2,"short sleeve"):
            top+=48
        elif eq(top2,"sleeveless"):
            top+=96
    #################################
        if eq(top3,"box"):
            top+=0
        elif eq(top3,"crop"):
            top+=8
        elif eq(top3,"halter"):
            top+=16
        elif eq(top3,"hoody"):
            top+=24
        elif eq(top3,"mantoman"):
            top+=32
        elif eq(top3,"neat"):
            top+=40

    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@
            
    ##################################shirts
    elif eq(top1,"shirts"):
        top=150
        if eq(top2,"long"):
            top+=0
        elif eq(top2,"short"):
            top+=16
    #################################
        if eq(top3,"business"):
            top+=0
        elif eq(top3,"casual"):
            top+=8
    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@
    ##################################outer
    elif eq(top1,"outer") and eq(top2,"outer"):
        top=182
    #################################
        if eq(top3,"cardigun"):
            top+=0
        elif eq(top3,"coat"):
            top+=8
        elif eq(top3,"hoodyzipup"):
            top+=16
        elif eq(top3,"jacket"):
            top+=24
        elif eq(top3,"padding"):
            top+=32
        elif eq(top3,"windbreaker"):
            top+=40
    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@

    ##################################onepiece
    elif eq(top1, "onepiece"):
        top=230
        bottoms=5
    ##################################@@@@@@@@@@@@@@@
    #//////////////////////////////상의



    ######@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@style
    #top
    if eq(topstyle,"colorful"):
        tstyle=3
    elif eq(topstyle,"lucid"):
        tstyle=4
    elif eq(topstyle,"natural"):
        tstyle=5
    elif eq(topstyle,"stiff"):
        tstyle=6
    elif eq(topstyle,"strong"):
        tstyle=7
    elif eq(topstyle,"warm"):
        tstyle=8
    elif eq(topstyle,"zingzy"):
        tstyle=9

    #bottoms
    if eq(bottomsstyle,"colorful"):
        bstyle=3
    elif eq(bottomsstyle,"lucid"):
        bstyle=4
    elif eq(bottomsstyle,"natural"):
        bstyle=5
    elif eq(bottomsstyle,"stiff"):
        bstyle=6
    elif eq(bottomsstyle,"strong"):
        bstyle=7
    elif eq(bottomsstyle,"warm"):
        bstyle=8
    elif eq(bottomsstyle,"zingzy"):
        bstyle=9
    if eq(top1, "onepiece"):
        stylepoint=1
    else :
        stylepoint=style_excel.cell(tstyle, bstyle).value

    clothpoint=cloth_excel.cell(top,bottoms).value
    allpoint=stylepoint+clothpoint
    
    return allpoint

#//////////////////////////////////////////////////////////

# MySQL Connection 연결
conn = pymysql.connect(host=os.environ['AWS_DB_HOST'],port=3306, user=os.environ['AWS_DB_USER'],
                       passwd=os.environ['AWS_DB_PASSWORD'], db=os.environ['AWS_DB_DATABASE'], charset='utf8')

# Connection 으로부터 Cursor 생성
curs = conn.cursor()
 
# SQL문 순서대로 검색, 삭제, 삽입


sql1 = "select cloImg,category,subclass1,subclass2,subclass3,subclass4,Color from dcloset"
sql2 = "delete from gicloset"
sql3 = "insert into gicloset(cloImg,recom1,recom2,recom3,recom4,recom5,section,count) values (%s, %s, %s, %s, %s, %s, %s,%s)"
sql4= "select gimg,gCategorize,gtype,glength,gsubclass,pattern,gstyle from gimg_analysis"

curs.execute(sql2)
curs.execute(sql4)
grows = curs.fetchall()
curs.execute(sql1)

# 데이타 Fetch
rows = curs.fetchall()
for row1 in rows:
#db에 넣을 값
    savedata=[]
#대상이 상의일때
    if eq(row1[1],"top") and not(eq(row1[2],"onepiece")) :
        for row2 in rows:
            if eq(row2[1],"bottoms") and not(eq(row2[2],"onepiece")):
                savedata.append([row2[0],usedvm(row2[2],row2[3],row2[4],row2[5],row2[6],
                                                row1[2],row1[3],row1[4],row1[5],row1[6])])
#대상이 하의일때
    elif eq(row1[1],"bottoms") and not(eq(row1[2],"onepiece")):
        for row2 in rows:
            if eq(row2[1],"top") and not(eq(row2[2],"onepiece")):
                 savedata.append([row2[0],usedvm(row1[2],row1[3],row1[4],row1[5],row1[6],
                                                 row2[2],row2[3],row2[4],row2[5],row2[6])])
#대상이 원피스일때
    else :
        for row2 in rows:
            if eq(row2[2] ,"onepiece"):
                savedata.append([row2[0],9])
    usedata=sorted(savedata, key=lambda x:x[1],reverse=True)
    if len(usedata) > 4:
        curs.execute(sql3,(row1[0],usedata[0][0],usedata[1][0],usedata[2][0],usedata[3][0],usedata[4][0],int(0),int(len(usedata))))
    elif len(usedata) == 4:
        curs.execute(sql3,(row1[0],usedata[0][0],usedata[1][0],usedata[2][0],usedata[3][0],None,int(0),int(len(usedata))))
    elif len(usedata) == 3:
        curs.execute(sql3,(row1[0],usedata[0][0],usedata[1][0],usedata[2][0],None,None,int(0),int(len(usedata))))
    elif len(usedata) == 2:
        curs.execute(sql3,(row1[0],usedata[0][0],usedata[1][0],None,None,None,int(0),int(len(usedata))))
    elif len(usedata) == 1:
        curs.execute(sql3,(row1[0],usedata[0][0],None,None,None,None,int(0),int(len(usedata))))
    else :
        curs.execute(sql3,(row1[0],None,None,None,None,None,int(0),int(len(usedata))))

#안경으로찍은 이미지+옷장
for grow in grows:
#db에 넣을 값
    savedata1=[]
#대상이 상의일때
    if eq(grow[1],"top") and not(eq(grow[2],"onepiece")) :
        for row2 in rows:
            if eq(row2[1],"bottoms") and not(eq(row2[2],"onepiece")):
                savedata1.append([row2[0],usedvm(row2[2],row2[3],row2[4],row2[5],row2[6],
                                                grow[2],grow[3],grow[4],grow[5],grow[6])])
#대상이 하의일때
    elif eq(grow[1],"bottoms") and not(eq(grow[2],"onepiece")):
        for row2 in rows:
            if eq(row2[1],"top") and not(eq(row2[2],"onepiece")):
                 savedata1.append([row2[0],usedvm(grow[2],grow[3],grow[4],grow[5],grow[6],
                                                 row2[2],row2[3],row2[4],row2[5],row2[6])])
#대상이 원피스일때
    else :
        for row2 in rows:
            if eq(row2[2],"onepiece"):
                savedata1.append([row2[0],9])
    usedata1=sorted(savedata1, key=lambda x:x[1],reverse=True)
    if len(usedata1) > 4:
        curs.execute(sql3,(grow[0],usedata1[0][0],usedata1[1][0],usedata1[2][0],usedata1[3][0],usedata1[4][0],int(1),int(len(usedata1))))
    elif len(usedata1) == 4:
        curs.execute(sql3,(grow[0],usedata1[0][0],usedata1[1][0],usedata1[2][0],usedata1[3][0],None,int(1),int(len(usedata1))))
    elif len(usedata1) == 3:
        curs.execute(sql3,(grow[0],usedata1[0][0],usedata1[1][0],usedata1[2][0],None,None,int(1),int(len(usedata1))))
    elif len(usedata1) == 2:
        curs.execute(sql3,(grow[0],usedata1[0][0],usedata1[1][0],None,None,None,int(1),int(len(usedata1))))
    elif len(usedata1) == 1:
        curs.execute(sql3,(grow[0],usedata1[0][0],None,None,None,None,int(1),int(len(usedata1))))
    else :
        curs.execute(sql3,(grow[0],None,None,None,None,None,int(1),int(len(usedata1))))
# DB에 Complete 하기
conn.commit()

# Connection 닫기
conn.close()
