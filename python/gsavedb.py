from openpyxl import load_workbook
from operator import eq
import pymysql.cursors

#엑셀 불러오기
load_cloth=load_workbook("vm.xlsx", data_only=True)
cloth_excel=load_cloth['cloth']
style_excel=load_cloth['style']

'''
설명 : 상의 하의의 정보를 받고 vm 엑셀 파일을 참조하여 코디 점수를 내주는 함수
인자 : Bottom 정보 5개, top 정보 5개
retrun값 : allpoint 합산 점수 (int)
'''

#사용자정의함수 vm사용
def usedvm(bottoms1,bottoms2,bottoms3,bottoms4,bottomsstyle,top1,top2,top3,top4,topstyle) :
    #//////////////////////////////하의
    ###############################pants
    if eq(bottoms1,"pants"):
        bottoms=5
    ###############################
        if eq(bottoms2,"long"):
            bottoms+=0
        elif eq(bottoms2,"short"):
            bottoms+=40
    ###############################
        if eq(bottoms3,"bootcut"):
            bottoms+=0
        elif eq(bottoms3, "cagopants"):
            bottoms+=8
        elif eq(bottoms3, "jeans"):
            bottoms+=16
        elif eq(bottoms3, "jogger"):
            bottoms+=24
        elif eq(bottoms3, "leggings"):
            bottoms+=32
    #################################
        if eq(bottoms4,"printing"):
            bottoms+=0
        elif eq(bottoms4, "dot"):
            bottoms+=1
        elif eq(bottoms4, "leopard"):
            bottoms+=2
        elif eq(bottoms4, "race"):
            bottoms+=3
        elif eq(bottoms4, "basic"):
            bottoms+=4
        elif eq(bottoms4, "stripe"):
            bottoms+=5
        elif eq(bottoms4, "check"):
            bottoms+=6
        elif eq(bottoms4, "camouflage"):
            bottoms+=7
    ###################################@@@@@@@@@@@@@@@@
            
    ###################################skirt
    elif eq(bottoms1,"skirt") and eq(bottoms2,"skirt"):
        bottoms=85
    ###############################
        if eq(bottoms3,"accordion skirt"):
            bottoms+=0
        elif eq(bottoms3, "a-line skirt"):
            bottoms+=8
        elif eq(bottoms3, "mini skirt"):
            bottoms+=16
        elif eq(bottoms3, "tube skirt"):
            bottoms+=24
        elif eq(bottoms3, "wrap skirt"):
            bottoms+=32
    #################################
        if eq(bottoms4,"printing"):
            bottoms+=0
        elif eq(bottoms4, "dot"):
            bottoms+=1
        elif eq(bottoms4, "leopard"):
            bottoms+=2
        elif eq(bottoms4, "race"):
            bottoms+=3
        elif eq(bottoms4, "basic"):
            bottoms+=4
        elif eq(bottoms4, "stripe"):
            bottoms+=5
        elif eq(bottoms4, "check"):
            bottoms+=6
        elif eq(bottoms4, "camouflage"):
            bottoms+=7

    ###################################@@@@@@@@@@@@@@@@
    #//////////////////////////////하의


            
    #//////////////////////////////상의
    ##################################t-shirts
    if eq(top1,"t-shirts"):
        top=6
        if eq(top2,"logn sleeve"):
            top+=0
        elif eq(top2,"short sleeve"):
            top+=48
        elif eq(top2,"sleeveless"):
            top+=96
    #################################
        if eq(top3,"box"):
            top+=0
        elif eq(top3,"crop"):
            top+=8
        elif eq(top3,"halter"):
            top+=16
        elif eq(top3,"hoody"):
            top+=24
        elif eq(top3,"mantoman"):
            top+=32
        elif eq(top3,"neat"):
            top+=40

    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@
            
    ##################################shirts
    elif eq(top1,"shirts"):
        top=150
        if eq(top2,"long"):
            top+=0
        elif eq(top2,"short"):
            top+=16
    #################################
        if eq(top3,"business"):
            top+=0
        elif eq(top3,"casual"):
            top+=8
    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@
    ##################################outer
    elif eq(top1,"outer") and eq(top2,"outer"):
        top=182
    #################################
        if eq(top3,"cardigun"):
            top+=0
        elif eq(top3,"coat"):
            top+=8
        elif eq(top3,"hoodyzipup"):
            top+=16
        elif eq(top3,"jacket"):
            top+=24
        elif eq(top3,"padding"):
            top+=32
        elif eq(top3,"windbreaker"):
            top+=40
    #################################
        if eq(top4,"printing"):
            top+=0
        elif eq(top4,"dot"):
            top+=1
        elif eq(top4,"leopard"):
            top+=2
        elif eq(top4,"race"):
            top+=3
        elif eq(top4,"basic"):
            top+=4
        elif eq(top4,"stripe"):
            top+=5
        elif eq(top4,"check"):
            top+=6
        elif eq(top4,"camouflage"):
            top+=7
    ##################################@@@@@@@@@@@@@@@@

    ##################################onepiece
    elif eq(top1, "onepiece"):
        top=230
        bottoms=5
    ##################################@@@@@@@@@@@@@@@
    #//////////////////////////////상의



    ######@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@style
    #top
    if eq(topstyle,"colorful"):
        tstyle=3
    elif eq(topstyle,"lucid"):
        tstyle=4
    elif eq(topstyle,"natural"):
        tstyle=5
    elif eq(topstyle,"stiff"):
        tstyle=6
    elif eq(topstyle,"strong"):
        tstyle=7
    elif eq(topstyle,"warm"):
        tstyle=8
    elif eq(topstyle,"zingzy"):
        tstyle=9

    #bottoms
    if eq(bottomsstyle,"colorful"):
        bstyle=3
    elif eq(bottomsstyle,"lucid"):
        bstyle=4
    elif eq(bottomsstyle,"natural"):
        bstyle=5
    elif eq(bottomsstyle,"stiff"):
        bstyle=6
    elif eq(bottomsstyle,"strong"):
        bstyle=7
    elif eq(bottomsstyle,"warm"):
        bstyle=8
    elif eq(bottomsstyle,"zingzy"):
        bstyle=9
    if eq(top1, "onepiece"):
        stylepoint=1
    else :
        stylepoint=style_excel.cell(tstyle, bstyle).value

    clothpoint=cloth_excel.cell(top,bottoms).value
    allpoint=stylepoint+clothpoint
    
    return allpoint

#//////////////////////////////////////////////////////////

# MySQL Connection 연결
conn = pymysql.connect(host='',
                       port=3306, user='', passwd='', db='Closet', charset='utf8')

# Connection 으로부터 Cursor 생성
curs = conn.cursor()
 
# SQL문 순서대로 검색, 삭제, 삽입
#dcloset 대상 ('사용자 옷') 정보 불러오기
sql1 = "select cloImg,category,subclass1,subclass2,subclass3,subclass4,Color from dcloset"
#새로운 정보 생성시마다 중복을 피하기 위해 모두 삭제후 재입력
sql2 = "delete from gicloset"
#대상인 옷1 | 어울리는 옷 1, 2, 3, 4, 5, section : 안경이 대상인지, 옷장 내의 옷이 대상인지 구분 /하여 insert
sql3 = "insert into gicloset(cloImg,recom1,recom2,recom3,recom4,recom5,section) values (%s, %s, %s, %s, %s, %s,%s)"
#gimg_analysis대상 '안경'으로 찍은 옷 정보 불러오기
sql4= "select gimg,gCategorize,gtype,glength,gsubclass,pattern,gstyle from gimg_analysis"

curs.execute(sql2)
curs.execute(sql4)
#안경 이미지 정보 가져오기 실행! (sql4)
grows = curs.fetchall()
curs.execute(sql1)

# 데이타 Fetch 
rows = curs.fetchall()
'''
rows =(사용자의 옷) 레코드가 담긴 변수
savedata[] : (옷 PK, 옷 코디) 구성으로 값을 리스트로 저장 후 정렬
'''

for row1 in rows:
#db에 넣을 값
    savedata=[]
#대상이 상의일때
    'row1 = top임  /row1[1]이 top 이고 onepiece가 아닐 때'
    if eq(row1[1],"top") and not(eq(row1[2],"onepiece")) :
        for row2 in rows:
            'row2 = bottom인 옷들!'
            if eq(row2[1],"bottoms") and not(eq(row2[2],"onepiece")):
                'row2의 PK, row2의 바지와 row 1의 코디 매칭값을 savedata에 저장'
                savedata.append([row2[0],usedvm(row2[2],row2[3],row2[4],row2[5],row2[6],
                                                row1[2],row1[3],row1[4],row1[5],row1[6])])
#대상이 하의일때
    elif eq(row1[1],"bottoms") and not(eq(row1[2],"onepiece")):
        for row2 in rows:
            if eq(row2[1],"top") and not(eq(row2[2],"onepiece")):
                 savedata.append([row2[0],usedvm(row1[2],row1[3],row1[4],row1[5],row1[6],
                                                 row2[2],row2[3],row2[4],row2[5],row2[6])])
#대상이 원피스일때
    else :
        for row2 in rows:
            if eq(row2[2],"onepiece"):
                savedata.append([row2[0],9])
                
    #savedata 내림차순으로 정렬            
    usedata=sorted(savedata, key=lambda x:x[1],reverse=True)
    #gicloset에 대상옷1, 어울리는 옷 1, 2, 3, 4, 5 insert
    curs.execute(sql3,(row1[0],usedata[0][0],usedata[1][0],usedata[2][0],usedata[3][0],usedata[4][0],0))


#안경으로찍은 이미지+옷장
for grow in grows:
#db에 넣을 값
    savedata1=[]
#대상이 상의일때
    if eq(grow[1],"top") and not(eq(grow[2],"onepiece")) :
        for row2 in rows:
            if eq(row2[1],"bottoms") and not(eq(row2[2],"onepiece")):
                savedata1.append([row2[0],usedvm(row2[2],row2[3],row2[4],row2[5],row2[6],
                                                grow[2],grow[3],grow[4],grow[5],grow[6])])
#대상이 하의일때
    elif eq(grow[1],"bottoms") and not(eq(grow[2],"onepiece")):
        for row2 in rows:
            if eq(row2[1],"top") and not(eq(row2[2],"onepiece")):
                 savedata1.append([row2[0],usedvm(grow[2],grow[3],grow[4],grow[5],grow[6],
                                                 row2[2],row2[3],row2[4],row2[5],row2[6])])
#대상이 원피스일때
    else :
        for row2 in rows:
            if eq(row2[2],"onepiece"):
                savedata1.append([row2[0],9])
    usedata1=sorted(savedata1, key=lambda x:x[1],reverse=True)
    curs.execute(sql3,(grow[0],usedata1[0][0],usedata1[1][0],usedata1[2][0],usedata1[3][0],usedata1[4][0],1))    

# DB에 Complete 하기
    conn.commit()

# Connection 닫기
conn.close()
